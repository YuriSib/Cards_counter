import openpyxl
from openpyxl import load_workbook
import os


def line_breaks(dir_path_):
    files = os.listdir(dir_path_)
    list_tables = [file for file in files if ".xlsx" in file]
    for table in list_tables:
        wb = openpyxl.load_workbook(f'{dir_path_}/{table}')
        ws = wb.active

        col_num = -1
        have_to_desc = False
        for col in ws.iter_cols(min_row=1, max_row=1, values_only=True):
            col_num += 1
            try:
                if 'писание' in col[0]:
                    have_to_desc = True
                    break
            except TypeError:
                print(f'Программа не смогла прочитать файл {table}')
                break
        if not have_to_desc:
            continue

        max_row = ws.max_row
        for row in range(2, max_row+1):
            old_value = ws.cell(row=row, column=col_num+1).value

            if not old_value:
                continue
            new_value = old_value.replace('•', ' <br> •').replace('<br>  <br> •', ' <br> •')

            letter_idx = 0
            for letter in new_value:
                if letter == ' ' or letter == '<' or letter == 'b' or letter == 'r' or letter == '>':
                    letter_idx += 1
                else:
                    break
            new_value = new_value[letter_idx:]
            # new_value = old_value[1:]
            ws.cell(row=row, column=col_num+1).value = new_value

        wb.save(rf'{dir_path_}\{table}')


def find_column_num(name, sheet):
    for col_names in sheet.iter_rows(min_row=1, max_row=1, min_col=0, max_col=300):
        col_num = 0
        for col in col_names:
            col_num += 1
            column_name = col.value[:-1] if col.value and col.value[-1] == ' ' else col.value
            if name == column_name:
                return col_num


def cell_replace(col_num, column_name, sheet, book, table):
    incorrect_list = ['Ссылка на фото', 'Код', 'Название', 'Код товар', 'код']
    incorrect_value = {
        'Ссылка на фото': 'Ссылки на фото', 'Код': 'Код товара',
        'Название': 'Наименование', 'Код товар': 'Код товара', 'код': 'Код товара'
    }
    if column_name in incorrect_list:
        name_new = incorrect_value[column_name]
        sheet.cell(row=1, column=col_num).value = name_new
        book.save(fr'C:\Users\Administrator\Desktop\Таблицы\Готовые к загрузке\{table}')
        return name_new
    return column_name


def union_table(dir_path_):
    union_table_path = 'Объединение таблиц.xlsx'
    files = os.path.exists(rf'{dir_path_}\{union_table_path}')
    if not files:
        union_wb = openpyxl.Workbook()
    else:
        union_wb = load_workbook(rf'{dir_path_}\{union_table_path}')
    union_ws = union_wb.active

    files = os.listdir(dir_path_)
    list_tables = [file for file in files if ".xlsx" in file and 'копия' not in file]

    for table in list_tables:
        row_to_start = union_ws.max_row + 1
        union_table_cols_list = [col[0] for col in union_ws.iter_cols(min_row=1, max_row=1, values_only=True) if col[0]]
        quantity_cols = len(union_table_cols_list)+1

        wb = openpyxl.load_workbook(f'{dir_path_}/{table}')
        ws = wb.active
        curr_table_cols_list = []
        col_num = 0
        for col in ws.iter_cols(min_row=1, max_row=1, values_only=True):
            col_num += 1
            if col[0]:
                column_name = col[0]
                if column_name[-1] == ' ':
                    column_name = column_name[:-1]
                column_name = cell_replace(col_num, column_name, ws, wb, table)
                # if column_name == 'Ссылка на фото':
                #     ws.cell(row=1, column=col_num).value = 'Ссылки на фото'
                #     wb.save(f'{dir_path_}/{table}')
                #     column_name = 'Ссылки на фото'
                curr_table_cols_list.append(column_name)

        for column_name in curr_table_cols_list:
            if not column_name:
                continue
            if column_name[-1] == ' ':
                column_name = column_name[:-1]
            if column_name in union_table_cols_list:
                continue
            else:
                union_ws.cell(row=1, column=quantity_cols).value = column_name
                union_table_cols_list.append(column_name)
                quantity_cols += 1

        for column_name in union_table_cols_list:
            if column_name in curr_table_cols_list:
                column_num = union_table_cols_list.index(column_name)+1

                curr_row = row_to_start
                curr_table_col_num = find_column_num(name=column_name, sheet=ws)
                if curr_table_col_num:
                    for values in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=curr_table_col_num, max_col=curr_table_col_num):
                        for value in values:
                            if value.value:
                                union_ws.cell(row=curr_row, column=column_num).value = value.value
                            curr_row += 1

        union_wb.save(rf'{dir_path_}\{union_table_path}')


if __name__ == "__main__":
    line_breaks(r'C:\Users\Administrator\Desktop\Таблицы\Готовые к загрузке')
    union_table(r'C:\Users\Administrator\Desktop\Таблицы\Готовые к загрузке')
