import os
import openpyxl
from datetime import datetime


def sorted_date(path_to_dir, path_to_table, wb, ws):
    date_format = "%d.%m.%y"
    # получаем список .xlsx документов
    table_list = [table for table in os.listdir(path_to_dir) if '.xlsx' in table and '~$' not in table]

    list_remote_work = [table for table in table_list if 'Д' in table]
    clean_table_list = [table.replace("Д", '') for table in table_list]

    # получаем список кортежей вида [(таблица.xlsx, день.месяц.год), ...,  (..., ...)]
    date_table_list_ = [(table, table.split('(')[1].replace("Д", '').replace(").xlsx", '').replace(" ", ''))
                        for table in clean_table_list]
    # получаем список дат вида [день.месяц.год1, ..., день.месяц.годN]
    date_string_list = [table.split('(')[1].replace(").xlsx", '').replace(" ", '') for table in clean_table_list]
    # приводим date_string_list к списку объектов datetime
    date_list = [datetime.strptime(date_string, date_format).date() for date_string in date_string_list]

    sorted_date_list = sorted(date_list)
    sorted_date_str_list = [date.strftime(date_format) for date in sorted_date_list]
    sorted_unique_date_list = []
    for item in sorted_date_str_list:
        if item not in sorted_unique_date_list:
            sorted_unique_date_list.append(item)

    date_row = 1
    for date in sorted_unique_date_list:
        date_row += 1
        ws.cell(row=date_row, column=3).value = date
    wb.save(path_to_table)
    return dict(date_table_list_), sorted_unique_date_list


def card_count(ws, table):
    column_num = 1
    for column in range(1, 10):
        desc_column = ws.cell(row=1, column=column).value
        try:
            if 'писание' in desc_column:
                column_num = column
                break
        except TypeError as e:
            print(f'Ошибка {e} возникла в {table}')

    num_rows = ws.max_row

    row_num = 0
    for row in range(2, num_rows):
        desc_row = ws.cell(row=row, column=column_num).value
        if desc_row:
            row_num += 1
    return row_num + 1


def table_counter(path, employee_):
    accounting_table = fr'C:\Users\Administrator\Desktop\Таблицы\Учет карточек {employee_}.xlsx'
    table_list = [table for table in os.listdir(path) if '.xlsx' in table and '~$' not in table]

    wb_cnt = openpyxl.load_workbook(accounting_table)
    ws_cnt = wb_cnt.active

    num_rows_in_cnt = ws_cnt.max_row
    # category_list = [ws_cnt.cell(row=row+1, column=1).value for row in range(1, num_rows_in_cnt)]

    date_table_dict, date_list = sorted_date(path_to_dir=path, path_to_table=accounting_table, wb=wb_cnt, ws=ws_cnt)

    for table in table_list:
        path_to_table = rf'{path}\{table}'
        try:
            wb = openpyxl.load_workbook(path_to_table)
        except ValueError:
            print(f'Не удалось прочитать таблицу {table}')
            continue
        ws = wb.active

        rows_num = card_count(ws, table)
        # clean_table_list = [table.replace("Д", '') for table in table_list]
        row_in_accounting_table = date_list.index(table.replace("Д", '').split('(')[1].replace(").xlsx", '').replace(" ", '')) + 2

        cell = ws_cnt.cell(row=row_in_accounting_table, column=2).value
        if cell:
            ws_cnt.cell(row=row_in_accounting_table, column=2).value = cell + rows_num
        else:
            ws_cnt.cell(row=row_in_accounting_table, column=2).value = rows_num

        category_name = table.split('(')[0]
        cell = ws_cnt.cell(row=row_in_accounting_table, column=1).value
        if cell:
            ws_cnt.cell(row=row_in_accounting_table, column=1).value = cell + ', ' + category_name
        else:
            ws_cnt.cell(row=row_in_accounting_table, column=1).value = category_name
        wb_cnt.save(accounting_table)


if __name__ == "__main__":
    list_employees = ['Вика', 'Олеся']
    for employee in list_employees:
        path_to_cards = fr'C:\Users\Administrator\Desktop\Общая папа\{employee}'
        table_counter(path_to_cards, employee)

