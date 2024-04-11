import os
import openpyxl


def change_name(big_table_path_, work_table_):
    big_wb = openpyxl.load_workbook(big_table_path_)
    big_ws = big_wb.active

    product_dict = {}
    for row in big_ws.iter_rows(min_col=1, max_col=2):
        product_dict[row[0].value] = row[1].value

    change_name_wb = openpyxl.load_workbook(work_table_)
    change_name_ws = change_name_wb.active

    row_num = 2
    for row in change_name_ws.iter_rows(min_row=2, min_col=1, max_col=1):
        clear_split_name = row[0].value.split('Изменено короткое название с ')[0].replace('Изменено название с ', '')\
            .replace('"', '').split(' на ')

        old_name, new_name = clear_split_name[0], clear_split_name[1]

        change_name_ws.cell(row=row_num, column=1).value = old_name
        change_name_ws.cell(row=row_num, column=2).value = new_name

        if old_name in product_dict:
            article = product_dict[old_name]
            change_name_ws.cell(row=row_num, column=3).value = article

        print(clear_split_name)
        row_num += 1

    change_name_wb.save(work_table_)


if __name__ == "__main__":
    big_table_path = r'C:\Users\Administrator\Desktop\Таблицы\Готовые к загрузке\Объединение таблиц.xlsx'
    work_table = r'C:\Users\Administrator\Desktop\Таблицы\Замена названий.xlsx'

    change_name(big_table_path, work_table)
