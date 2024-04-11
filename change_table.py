import openpyxl
from openpyxl import load_workbook
import os


PRICE_NAME = [
    'Прайс 12.06.21 старое 9 января',
    'Прайс 12.06.21 ул 9 января',
    'Прайс 18.04.22 Хмелева',
    'Прайс 20.09.21 солнечный',
    'Прайс Молодогвардеец 09.06.22',
]


def work_with_column(path_, table_, right_price_, prises):
    wb = openpyxl.load_workbook(f'{path_}/{table_}')
    ws = wb.active

    prises.remove(right_price_)

    col_range = [row.value for row in ws[1]]

    right_price_column = col_range.index(right_price_) + 1
    for_change_1, for_change_2, for_change_3, for_change_4 = \
        col_range.index(prises[0]) + 1, col_range.index(prises[1]) + 1, col_range.index(prises[2]) + 1, \
        col_range.index(prises[3]) + 1
    code = col_range.index('Код') + 1

    row_num = 1
    for row in ws.iter_rows(min_row=2, min_col=right_price_column, max_col=right_price_column):
        row_num += 1

        curr_code = ws.cell(row=row_num, column=code).value
        if not curr_code:
            continue

        if not row[0].value:
            for column in [for_change_1, for_change_2, for_change_3, for_change_4]:
                change_value = ws.cell(row=row_num, column=column).value
                if change_value:
                    break

            ws.cell(row=row_num, column=right_price_column).value = change_value
    wb.save(f'{path_}/{table_}')

    return col_range


if __name__ == "__main__":
    path = r'C:\Users\Administrator\Desktop\Таблицы\Для изменения цен каталога'
    table = 'Каталог ч5.xlsx'
    right_price = 'Прайс Молодогвардеец 09.06.22'
    change_price = 'Прайс 12.06.21 старое 9 января'

    print(work_with_column(path, table, right_price, PRICE_NAME))

